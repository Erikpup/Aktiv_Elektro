import sqlite3

from kivy.core.window import Window
from kivy.metrics import dp
from kivy.uix.button import Button
from kivy.uix.checkbox import CheckBox
from kivy.uix.dropdown import DropDown
from kivy.uix.gridlayout import GridLayout
from kivy.uix.label import Label
from kivy.uix.popup import Popup
from kivy.uix.recycleview import RecycleView
from kivy.uix.screenmanager import Screen
from kivy.uix.textinput import TextInput


class Material():
    def __init__(self, **kw):
        self.file_items = 'itemsbase.db'

    def items_open(self, selff, key_list, button_key):
        selff.title.text = button_key
        selff.title.font_size = 80
        selff.back1.text = '<'
        selff.add.on_press = selff.add_work
        selff.add.text = '+'
        selff.sr.clear_widgets()
        selff.sr.size_hint = (1, .95)
        selff.sr.pos_hint = {'center_x': .5, 'center_y': .58}
        layout = GridLayout(cols=4, spacing=10, size_hint=(1, None),
                                 pos_hint={'center_x': .5, 'center_y': .5})
        layout.bind(minimum_height=layout.setter('height'))
        root = RecycleView(size_hint=(1, .8))
        root.add_widget(layout)
        sql = "SELECT * FROM albums"
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        info = cursor1.fetchall()
        keys_move = {}
        i = 0
        for info_btn in info:
            def redc(a):
                layout = Screen()
                conn = sqlite3.connect("items.db")
                cursor = conn.cursor()
                font = Window.size[0] * Window.size[1] * .000025
                self.popup = Popup(title='     Изменение',
                                   content=layout, size_hint=(.8, .6))
                button_back = Button(text='Отменить', size_hint=(.5, .2), pos_hint={'center_x': .25, 'center_y': .1},
                                     font_size=font)
                button_save = Button(text='Сохранить', size_hint=(.5, .2), pos_hint={'center_x': .75, 'center_y': .1},
                                     font_size=font)
                input_name = TextInput(text=f'', size_hint=(.5, .13),
                                       pos_hint={'center_x': .325, 'center_y': .55},
                                       font_size=font)
                info_list = [['Проектный ввод', 'Ручной ввод']]
                y_x = .85
                for i in info_list:
                    label1 = Label(text=i[0], size_hint=(.25, .15), pos_hint={'center_x': .15, 'center_y': y_x},
                                   font_size=font * .8)
                    label2 = Label(text=i[1], size_hint=(.25, .15), pos_hint={'center_x': .65, 'center_y': y_x},
                                   font_size=font * .8)
                    layout.add_widget(label2)
                    layout.add_widget(label1)
                label2 = Label(text='Умножить', size_hint=(.25, .15), pos_hint={'center_x': .7, 'center_y': .55},
                               font_size=font * .8)
                layout.add_widget(label2)
                def check_2(a, b):
                    if b.active == True and a.active == True:
                        b.active = False
                    if b.active == False and a.active == False:
                        a.active = True

                kir = CheckBox(active=False, size_hint=(.1, .1), pos_hint={'center_x': .35, 'center_y': .85},
                               on_press=lambda a: check_2(kir, but))
                but = CheckBox(active=False, size_hint=(.1, .1), pos_hint={'center_x': .85, 'center_y': .85},
                               on_press=lambda a: check_2(but, kir))
                kir1 = CheckBox(active=False, size_hint=(.1, .1), pos_hint={'center_x': .85, 'center_y': .55},
                               on_press=lambda a: check_2(kir, but))

                sql = "SELECT * FROM put WHERE ID=?"
                conn1 = sqlite3.connect('items.db')
                cursor1 = conn1.cursor()
                cursor1.execute(sql, (str(keys_move[a][0]),))
                infoo = cursor1.fetchall()
                if infoo != []:
                    but.active = True
                    input_name.text = infoo[0][1]
                    if infoo[0][2] == 'y':
                        kir1.active =True
                else:
                    kir.active=True
                    input_name.text = '0'

                def ok_save():
                    if kir.active == True:
                        sql = "SELECT * FROM put WHERE ID=?"
                        conn1 = sqlite3.connect('items.db')
                        cursor1 = conn1.cursor()
                        cursor1.execute(sql, (str(keys_move[a][0]),))
                        if cursor1.fetchall() != []:
                            sql = "DELETE FROM put WHERE ID = ?"
                            cursor1.execute(sql, (str(keys_move[a][0]),))
                    if but.active ==True:
                        sql = "SELECT * FROM put WHERE ID=?"
                        conn1 = sqlite3.connect('items.db')
                        cursor1 = conn1.cursor()
                        cursor1.execute(sql, (str(keys_move[a][0]),))
                        if cursor1.fetchall() == []:
                            if kir1.active ==True:
                                form = 'y'
                            else:
                                form = 'n'
                            sql = f"""
                                                                    INSERT INTO put (ID,col,form)
                                                                    VALUES ('{(str(keys_move[a][0]))}','{input_name.text}','{form}');
                                                                    """
                            cursor1.execute(sql)
                        else:
                            if kir1.active ==True:
                                form = 'y'
                            else:
                                form = 'n'
                            sql = f"""
                                                                    UPDATE put
                                                                    SET col = '{input_name.text}' 
                                                                    WHERE ID = '{(str(keys_move[a][0]))}'
                                                                    """
                            cursor1.execute(sql)
                            conn1.commit()
                            sql = f"""
                                                                    UPDATE put
                                                                    SET form = '{form}'
                                                                    WHERE ID = '{(str(keys_move[a][0]))}'
                                                                    """
                            cursor1.execute(sql)
                    conn1.commit()
                    self.popup.dismiss()
                button_back.bind(on_press=lambda *args: self.popup.dismiss())
                layout.add_widget(button_back)
                button_save.bind(on_press=lambda *args: ok_save())
                layout.add_widget(button_save)
                layout.add_widget(input_name)
                layout.add_widget(kir)
                layout.add_widget(kir1)
                layout.add_widget(but)
                self.popup.open()

            sql = f"SELECT rowid, * FROM name_group WHERE id = '{button_key.split(' ')[0]}'"
            conn1 = sqlite3.connect('items.db')
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            inf = cursor1.fetchall()[0]
            true = False
            for f in inf[3].split(','):
                if str(info_btn[0]) == str(f):
                    true=True
                    continue

            if true == False:
                continue
            i += 1
            if len(info_btn[1]) >= 29:
                text = info_btn[1][:28] + '\n' + info_btn[1][28:]
            else:
                text = info_btn[1]
            btn = Button(text=f'{i}.{text}',
                              size_hint_y=None, size_hint_x=.8,
                              height=dp(40),
                              background_normal='button.png',
                              font_size=20,
                              color=[0, 0, 1, 1],
                              on_press=redc)
            layout.add_widget(btn)
            font = Window.size[0] * Window.size[1] * .000013
            input_size = TextInput(text=f'{(info_btn[2])}', size_hint=(.1, .065),
                                   pos_hint={'center_x': .85, 'center_y': .89},
                                   font_size=font)
            layout.add_widget(input_size)
            input_cost = TextInput(text=f'{(info_btn[3])}', size_hint=(.1, .065),
                                   pos_hint={'center_x': .85, 'center_y': .89},
                                   font_size=font)
            layout.add_widget(input_cost)
            redact = Button(text=f'Связка',
                                 size_hint_y=None, size_hint_x=.15,
                                 height=dp(40),
                                 background_normal='button.png',
                                 font_size=20,
                                 color=[0, 0, 1, 1],
                                 on_press=lambda a: selff.items_svizka(selff, a, keys_move,key_list, button_key))
            layout.add_widget(redact)
            keys_move.setdefault(redact, [info_btn[0],input_cost,input_size,info_btn[1]])
            keys_move.setdefault(btn, [info_btn[0],input_cost,input_size])
        selff.sr.add_widget(root)
        def save(a):
            for key in keys_move.keys():
                sql = f"SELECT * FROM albums WHERE ID = ?"
                conn1 = sqlite3.connect('items.db')
                cursor1 = conn1.cursor()
                cursor1.execute(sql,(int(keys_move[key][0]),))
                info_name = cursor1.fetchall()
                sql = f"""
                                            UPDATE albums
                                            SET pcs = '{keys_move[key][2].text}'
                                            WHERE ID = '{info_name[0][0]}'
                                            """
                cursor1.execute(sql)
                conn1.commit()
                sql = f"""
                                            UPDATE albums
                                            SET price = '{keys_move[key][1].text}'
                                            WHERE ID = {info_name[0][0]}
                                            """

                cursor1.execute(sql)
                conn1.commit()
        def redc():
            layout = Screen()
            conn = sqlite3.connect("items.db")
            cursor = conn.cursor()
            font = Window.size[0] * Window.size[1] * .000025
            self.popup = Popup(title='     Создание',
                               content=layout, size_hint=(.8, .6))
            button_back = Button(text='Отменить', size_hint=(.5, .2), pos_hint={'center_x': .25, 'center_y': .1},
                                 font_size=font)
            button_save = Button(text='Сохранить', size_hint=(.5, .2), pos_hint={'center_x': .75, 'center_y': .1},
                                 font_size=font)
            input_name = TextInput(text=f'', size_hint=(.65, .16),
                                   pos_hint={'center_x': .325, 'center_y': .85},
                                   font_size=font)
            button_back.bind(on_press=lambda *args: self.popup.dismiss())
            layout.add_widget(button_back)
            button_save.bind(on_press=lambda *args: ok_save())
            layout.add_widget(button_save)
            layout.add_widget(input_name)
            layout.add_widget(input_cost)
            self.popup.open()
        def add_items():
            layout = Screen()
            conn = sqlite3.connect("items.db")
            cursor = conn.cursor()
            font = Window.size[0] * Window.size[1] * .000025
            self.popup = Popup(title='     Создание',
                               content=layout, size_hint=(.8, .6))
            button_back = Button(text='Отменить', size_hint=(.5, .2), pos_hint={'center_x': .25, 'center_y': .1},
                                 font_size=font)
            button_save = Button(text='Сохранить', size_hint=(.5, .2), pos_hint={'center_x': .75, 'center_y': .1},
                                 font_size=font)
            input_name = TextInput(text=f'', size_hint=(.65, .16),
                                   pos_hint={'center_x': .325, 'center_y': .85},
                                   font_size=font)
            input_cost = TextInput(text=f'{0}', size_hint=(.25, .065),
                                   pos_hint={'center_x': .85, 'center_y': .89},
                                   font_size=font)
            list_list = [['шт.', "м/п"]]
            dropdown1 = DropDown()
            for index in list_list[0]:
                btn = Button(text=f'{index}', size_hint_y=None, height=60, font_size=font)
                btn.bind(on_release=lambda btn: dropdown1.select(btn.text))
                dropdown1.add_widget(btn)
            self.mainbutton = Button(color=(0, 0, 0, 1),
                                     background_color=[1, 1, 1, 100],
                                     background_normal='', text=f'{list_list[0][0]}',
                                     pos_hint={'center_x': .85, 'center_y': .81}, size_hint=(.25, .065), font_size=font)

            layout.add_widget(self.mainbutton)
            self.mainbutton.bind(on_release=dropdown1.open)

            dropdown1.bind(on_select=lambda instance, x: setattr(self.mainbutton, 'text', x))



            def ok_save(*args):
                sql = "SELECT * FROM albums"
                conn1 = sqlite3.connect('items.db')
                cursor11 = conn1.cursor()
                cursor11.execute(sql)
                info_cursor11 = cursor11.fetchall()
                max_id = int(info_cursor11[len(info_cursor11)-1][0])+1
                sql = f"""
                                                    INSERT INTO albums (ID,Name,pcs,price,work)
                                                    VALUES ({max_id},'{input_name.text}', '{self.mainbutton.text}',
                                                     '{input_cost.text}', '');
                                                    """
                cursor.execute(sql)
                conn.commit()

                sql = f"SELECT rowid, * FROM name_group WHERE id = '{button_key.split(' ')[0]}'"
                conn1 = sqlite3.connect('items.db')
                cursor1 = conn1.cursor()
                cursor1.execute(sql)
                info_g = cursor1.fetchall()[0]

                sql = F"UPDATE name_group SET info='{info_g[3] + ',' + str(max_id)}' " \
                      F"WHERE id = '{button_key.split(' ')[0]}'"
                cursor.execute(sql)
                conn.commit()
                self.popup.dismiss()
                selff.func_open_group(button_key, key_list)

            button_back.bind(on_press=lambda *args: self.popup.dismiss())
            layout.add_widget(button_back)
            button_save.bind(on_press=lambda *args: ok_save())
            layout.add_widget(button_save)
            layout.add_widget(input_name)
            layout.add_widget(input_cost)
            self.popup.open()

        selff.add.on_press = add_items
        button_add = Button(
            text='Сохранить',
            background_color=[0, 0, 1, 100],
            background_normal='',
            font_size=80,
            color=[1, 1, 1, 1],
            size_hint=(1, .1),
            pos_hint={'center_x': .5, 'center_y': -.06},
            on_press = save)
        def okey():
            sql = f"SELECT rowid, * FROM name_group "
            conn1 = sqlite3.connect('items.db')
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            inf = cursor1.fetchall()
            nfo =''
            for info_info in inf:
                if info_info[3] != None and button_key.split(' ')[0] in info_info[3].split(','):
                    nfo = info_info[1] + ' ' + info_info[2]
            selff.update_list_group(nfo,key_list )

        selff.back1.on_press = okey
        selff.sr.add_widget(button_add)

    def items_svizka(self,selff,button, list_info_button,key_list, button_key):
        selff.back1.text = '<'
        if len(list_info_button[button][3]) >= 29:
            selff.title.text = list_info_button[button][3][:28] + '\n' + list_info_button[button][3][28:]
        else:
            selff.title.text = list_info_button[button][3]
        selff.title.font_size=35
        selff.back1.on_press = lambda :self.items_open(selff,key_list, button_key)
        selff.add.on_press = selff.add_work
        selff.add.text = '+'
        selff.sr.clear_widgets()
        selff.sr.size_hint = (1, .95)
        selff.sr.pos_hint = {'center_x': .5, 'center_y': .58}
        layout = GridLayout(cols=2, spacing=1, size_hint=(1, None),
                            pos_hint={'center_x': .5, 'center_y': .5})
        layout.bind(minimum_height=layout.setter('height'))
        root = RecycleView(size_hint=(1, .8))
        root.add_widget(layout)
        sql = "SELECT * FROM albums "
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        info_items = cursor1.fetchall()
        if button_key.split('.')[0] != '2':
            sql = "SELECT * FROM projects "
            conn1 = sqlite3.connect('workdatabase.db')
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            info = cursor1.fetchall()
        else:
            sql = "SELECT * FROM works "
            conn1 = sqlite3.connect('noworkdatabase.db')
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            info = cursor1.fetchall()
        keys_move = {}
        i = 0
        for info_btn in info:
            i += 1
            text = info_btn[0]
            btn = Button(text=f'{i}.{text}',
                         size_hint_y=None, size_hint_x=1,
                         height=dp(40),
                         background_normal='button.png',
                         font_size=20,
                         color=[0, 0, 1, 1])
            layout.add_widget(btn)
            active = CheckBox(active=False, size_hint_x=.1)
            for item in info_items:
                if item[0] == int(list_info_button[button][0]):
                    if item[4] != None and text in item[4]:
                        active.active = True
            layout.add_widget(active)
            keys_move.setdefault(active,btn.text)
        selff.sr.add_widget(root)
        def save(button1):
            sql = f"SELECT * FROM albums WHERE ID = {int(list_info_button[button][0])}"
            conn1 = sqlite3.connect('items.db',)
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            info_name = cursor1.fetchall()
            list_work = ''
            for key in keys_move.keys():
                if key.active == True:
                    list_work += (keys_move[key]) + ','
            sql = f"""
                            UPDATE albums
                            SET work = '{list_work}'
                            WHERE ID = '{info_name[0][0]}'
                            """
            cursor1.execute(sql)
            conn1.commit()
            button_add.pos_hint={'center_x': 5, 'center_y': .05}
            self.items_open(selff,key_list, button_key)


        button_add = Button(
            text='Сохранить',
            background_color=[0, 0, 1, 100],
            background_normal='',
            font_size=80,
            color=[1, 1, 1, 1],
            size_hint=(1, .1),
            pos_hint={'center_x': .5, 'center_y': -.06},
            on_press=save)
        selff.sr.add_widget(button_add)

    def open_info_in_excel(self):
        import openpyxl
        import sqlite3
        import os
        from kivy.utils import platform
        if platform == "android":
            wookbook = openpyxl.load_workbook('/storage/emulated/0/folder/items.xlsx')
        else:
            wookbook = openpyxl.load_workbook('items.xlsx')
        worksheet = wookbook.active
        total = []
        total_group = []
        for i in range(0, worksheet.max_row):
            row = []
            for col in worksheet.iter_cols(1, worksheet.max_column-1):
                row.append(col[i].value)
            total.append(row)
        for i in range(0, worksheet.max_row):
            row = []
            for col in worksheet.iter_cols(1, worksheet.max_column):
                row.append(col[i].value)
            total_group.append(row)
        conn = sqlite3.connect("items.db")  # или :memory: чтобы сохранить в RAM
        cursor = conn.cursor()
        cursor.execute('''
                        DROP TABLE IF EXISTS albums ''')

        # Создание таблицы
        try:
            cursor.execute("""CREATE TABLE albums
                              (ID int, Name text, pcs text,
                               price text, work text)
                           """)
        except sqlite3.OperationalError:
            pass
        cursor.executemany("INSERT INTO albums VALUES (?,?,?,?,?)", total[1:len(total)])
        conn.commit()

        sql = "SELECT rowid, * FROM name_group"
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        info_g = cursor1.fetchall()
        for group in info_g:
            list_id_items = ''
            for item in total_group:
                if str(item[5].split(' ')[0]) == str(group[1]) :
                    list_id_items += str(int(item[0])) +','
            if list_id_items != '':
                sql = f"UPDATE name_group SET info='{list_id_items}' WHERE id='{str(group[1])}'"
                cursor1.execute(sql)
                conn1.commit()





        layout = Screen()
        popup = Popup(title='     Загрузка завершена',
                           content=layout, size_hint=(.5, .2),
                           pos_hint={"center_x": .5, "center_y": .45})

        layout.add_widget(
            Button(text="Готово", size_hint=(1, .5), on_press=lambda a: popup.dismiss(),
                   pos_hint={"center_x": .5}))
        popup.open()

    def create_excel(self):
        import sqlite3
        import pandas as pd

        conn = sqlite3.connect('items.db')
        cursor = conn.cursor()
        sql = "SELECT * FROM albums"
        cursor.execute(sql)
        info = cursor.fetchall()
        sql = "SELECT rowid, * FROM name_group"
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        info_g = cursor1.fetchall()
        cols = []
        for i in range(5):
            col = []
            for row in info:
                col.append(row[i])
            cols.append(col)
        col = []
        for i in info:
            for row in info_g:
                if str(i[0]) in row[3].split(','):
                    col.append(row[1]+' '+row[2])
                    break
        cols.append(col)
        df = pd.DataFrame({'ID': cols[0],
                           'Name': cols[1],
                           'pcs': cols[2],
                           'price': cols[3],
                           'works': cols[4],
                           'group':cols[5]})
        from kivy.utils import platform
        if platform == "android":
            writer = pd.ExcelWriter('/storage/emulated/0/folder/items.xlsx', engine='xlsxwriter')
        else:
            writer = pd.ExcelWriter('items.xlsx', engine='xlsxwriter')
        df.to_excel(writer, sheet_name='Group1', index=False)
        for column in df:
            column_width = max(df[column].astype(str).map(len).max(), len(column))
            col_idx = df.columns.get_loc(column)
            writer.sheets['Group1'].set_column(col_idx, col_idx, column_width)
        writer.save()
        layout = Screen()
        popup = Popup(title='     Загрузка завершена',
                           content=layout, size_hint=(.5, .2),
                           pos_hint={"center_x": .5, "center_y": .45})

        layout.add_widget(
            Button(text="Готово", size_hint=(1, .5), on_press=lambda a: popup.dismiss(),
                   pos_hint={"center_x": .5}))
        popup.open()

    def items_open_room(self, selff,j):
        def i():
            layout = Screen()
            button_back = Button(text='Нет', size_hint=(.5, .9), pos_hint={'center_x': .25, 'center_y': .5})
            button_save = Button(text='Да', size_hint=(.5, .9), pos_hint={'center_x': .75, 'center_y': .5})
            self.popup = Popup(title=f'          Вы уверены что хотите выйти?',
                               content=layout, size_hint=(.7, .25), pos_hint={'center_x': .5, 'center_y': .65})
            button_back.bind(on_press=self.popup.dismiss)


            def ok(a):
                conn = sqlite3.connect(selff.mydatabase)
                cursor = conn.cursor()
                sql = "SELECT rowid, * FROM projects WHERE rowid=?"
                cursor.execute(sql, (selff.button_text,))
                selff.func_info(cursor.fetchall())
                self.popup.dismiss()

            layout.add_widget(button_back)
            button_save.bind(on_press=ok)
            layout.add_widget(button_save)
            self.popup.open()

        selff.back.text = '<'
        selff.back.on_press = i
        selff.button_room = j

        def func_add():
            conn1 = sqlite3.connect(selff.roomsdatabase)
            cursor1 = conn1.cursor()
            cursor1.execute(sqlite_update_query, (str(selff.button_text),))
            conn1.commit()
            p = 0
            for key in list_total.keys():
                if self.keys_work[int(key[1])].text == '':
                    p = 1
                for v in self.keys_work[int(key[1])].text:
                    if (v in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']) == False:
                        p = 2
                if p == 1:
                    break

                if str(list_total[key][0]) == self.keys_work[int(key[1])].text and str(list_total[key][1]) == 'f':
                    block1 = 'f'
                else:
                    block1 = 't.n'

                sql = f"""
                                        INSERT INTO material (ID,address,col,block)
                                        VALUES (?, ?, ?, ?);
                                        """

                conn1 = sqlite3.connect(selff.roomsdatabase)
                cursor1 = conn1.cursor()
                cursor1.execute(sql, (str(key[1]), str(selff.button_text), self.keys_work[int(key[1])].text, block1,))
                conn1.commit()
                cursor1.execute(
                    f"SELECT rowid, * FROM material WHERE address={str(selff.button_text)} and ID={str(key[1])}")
            if p != 0:
                layout = Screen()
                button_save = Button(text='Okeй', size_hint=(.95, .9), pos_hint={'center_x': .5, 'center_y': .5})
                if p == 1:
                    self.popup = Popup(title=f'          Заполнены не все поля',
                                       content=layout, size_hint=(.7, .25), pos_hint={'center_x': .5, 'center_y': .65})
                else:
                    self.popup = Popup(title=f'          Введены лишние символы',
                                       content=layout, size_hint=(.7, .25), pos_hint={'center_x': .5, 'center_y': .65})

                def okt(a):
                    self.popup.dismiss()

                button_save.bind(on_press=okt)
                layout.add_widget(button_save)
                self.popup.open()
                return
            self.room_new = False
            selff.items_add_room(j)

        selff.add.on_press = func_add
        selff.sr.clear_widgets()
        selff.sr.size_hint = (1, .85)
        selff.sr.pos_hint = {'center_x': .5, 'center_y': .55}
        selff.title.text = 'Материал'
        selff.layout = GridLayout(cols=4, spacing=10, size_hint=(1, None),
                                 pos_hint={'center_x': .5, 'center_y': .5})
        selff.layout.bind(minimum_height=selff.layout.setter('height'))
        root = RecycleView(size_hint=(1, .9))
        root.add_widget(selff.layout)
        sql = "SELECT rowid, * FROM projects WHERE address=?"
        conn1 = sqlite3.connect(selff.roomsdatabase)
        cursor1 = conn1.cursor()
        cursor1.execute(sql, (str(selff.button_text),))
        information = cursor1.fetchall()
        sql = "SELECT rowid, * FROM material WHERE address=?"
        conn1 = sqlite3.connect(selff.roomsdatabase)
        cursor1 = conn1.cursor()
        cursor1.execute(sql, (str(selff.button_text),))
        information_col_material = cursor1.fetchall()
        keys_move = {}
        self.keys_work = {}
        conn = sqlite3.connect('items.db')
        cursor = conn.cursor()
        cursor.execute(f"SELECT rowid, * FROM albums")
        info_material = cursor.fetchall()
        conn = sqlite3.connect(selff.noworkdatabase)
        cursor = conn.cursor()
        cursor.execute(f"SELECT rowid, * FROM works")
        info131 = cursor.fetchall()
        str_work = {}
        str_avto = {}
        for info in information:
            if info[4] != None:
                for info_btn in info[4].split('-'):
                    if info_btn != '' and info_btn.split("*")[1] != '0':
                        if info_btn.split("*")[0] in str_work.keys():
                            str_work[info_btn.split("*")[0]]= str(int(info_btn.split("*")[1])+int(str_work[info_btn.split("*")[0]]))
                        else:
                            str_work.setdefault(info_btn.split("*")[0], str(info_btn.split("*")[1]))
            else:
                for info_btn in info[3].split('-'):
                    if info_btn != '' and info_btn.split("*")[1] != '0':
                        if info_btn.split("*")[0] in str_work.keys():
                            str_avto[info_btn.split("*")[0]] = str(
                                int(info_btn.split("*")[1]) + int(str_work[info_btn.split("*")[0]]))
                        else:
                            str_avto.setdefault(info_btn.split("*")[0], str(info_btn.split("*")[1]))




        list_material = {}
        sql = f"SELECT rowid, * FROM tip_avto WHERE adress ='{str(selff.button_text)}'"
        conn1 = sqlite3.connect(selff.roomsdatabase)
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        tip =cursor1.fetchall()
        sql = f"SELECT rowid, * FROM name_group"
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        for name_group in cursor1.fetchall():
            if len(name_group[3].split(',')[0].split('.'))==1:
                if (tip[0][2].split(' ')[0].split('.')[0] == name_group[1].split('.')[0] and tip[0][2].split(' ')[0].split('.')[1] == name_group[1].split('.')[1] and tip[0][2].split(' ')[0].split('.')[2] == name_group[1].split('.')[2]) or '1' in name_group[1].split('.')[0]:
                    for id_item in name_group[3].split(','):
                        if id_item == '':
                            continue
                        sql = f"SELECT rowid, * FROM albums WHERE ID = ?"
                        conn1 = sqlite3.connect('items.db')
                        cursor1 = conn1.cursor()
                        cursor1.execute(sql,(int(id_item),))
                        name_material=cursor1.fetchall()
                        if name_material ==[]:
                            continue
                        name_material=name_material[0]

                        if name_material[5] != None and name_material[5] != '':
                            for y in range(len(name_material[5].split(','))):
                                if '1' in name_group[1].split('.')[0]:
                                    if name_material[5].split(',')[y].split('.')[0] != '' and not(name_material in list_material) \
                                            and name_material[5].split(',')[y].split('.')[0] in str_work.keys():
                                        list_material.setdefault(name_material, 0)
                                    if name_material[5].split(',')[y].split('.')[0] != '' and name_material[5].split(',')[y].split('.')[0] in str_work.keys():
                                        list_material[name_material] += int(str_work[name_material[5].split(',')[y].split('.')[0]])
                                else:
                                    if name_material[5].split(',')[y].split('.')[0] != '' and not(name_material in list_material) \
                                            and name_material[5].split(',')[y].split('.')[0] in str_avto.keys():
                                        list_material.setdefault(name_material, 0)
                                    if name_material[5].split(',')[y].split('.')[0] != '' and name_material[5].split(',')[y].split('.')[0] in str_avto.keys():
                                        list_material[name_material] += int(str_avto[name_material[5].split(',')[y].split('.')[0]])

        for keys in list_material.keys():
            sql = f"SELECT rowid, * FROM put WHERE ID='{keys[1]}'"
            conn1 = sqlite3.connect('items.db')
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            infoo=cursor1.fetchall()
            if infoo != []:
                if infoo[0][3] == 'y':
                    list_material[keys] = list_material[keys] * int(infoo[0][2])
                else:
                    list_material[keys] = int(infoo[0][2])



        true_info={}
        for item_room in information_col_material:
            if item_room[4] !='f':
                for name_material in info_material:
                    if str(name_material[1])==item_room[1]:
                        true_info.setdefault(name_material,[item_room[3],item_room[4]])


        list_total = {}
        list_total_0 = {}
        for iol in info_material:
            true_false = False
            for key in true_info.keys():
                if key == iol and true_info[key][0] != '0':
                    list_total.setdefault(key, [true_info[key][0], true_info[key][1]])
                    true_false=True
                elif key == iol and true_info[key][0] == '0':
                    list_total_0.setdefault(key, [true_info[key][0], true_info[key][1]])
                    true_false=True
            if true_false == False:
                for p in list_material.keys():
                    if p == iol:
                        if iol[1] == 2:
                            print(iol[1])
                        if true_info != []:
                            list_total.setdefault(p,[list_material[p],'f'])

                        else:
                            list_total.setdefault(p,[list_material[p],'f'])
        for info in list_total_0.keys():
            list_total.setdefault(info,[list_total_0[info][0],list_total_0[info][1]])

        # print(list_material,'\n',str_total)
        sql = f"""
                    INSERT INTO material (ID,address,col,block)
                    VALUES (?, ?, ?, ?);
                    """
        sqlite_update_query = """DELETE FROM material WHERE address = ?"""

        conn1 = sqlite3.connect(selff.roomsdatabase)
        cursor1 = conn1.cursor()
        cursor1.execute(sqlite_update_query, (str(selff.button_text),))
        conn1.commit()
        for key in list_total.keys():
            # if list_total[key][0] == '0':
            cursor1.execute(sql, (str(key[1]),str(selff.button_text),list_total[key][0],list_total[key][1],))
            conn1.commit()



        for info_btn in list_total.keys():
            def block_func(item_button):
                layout = Screen()
                button_back = Button(text='Нет', size_hint=(.5, .3), pos_hint={'center_x': .25, 'center_y': 1})
                button_save = Button(text='Да', size_hint=(.5, .3), pos_hint={'center_x': .75, 'center_y': 1})
                self.popup = Popup(title=f'          Обновить данные {item_button.text}',
                                   content=layout, size_hint=(.7, .25), pos_hint={'center_x': .5, 'center_y': .65})
                button_back.bind(on_press=self.popup.dismiss)
                def okey_block(a):
                    sql = "UPDATE material SET block=? WHERE address=? and ID = ?"
                    conn1 = sqlite3.connect(selff.roomsdatabase)
                    cursor1 = conn1.cursor()
                    cursor1.execute(sql, ('f',str(selff.button_text),item_button.text.split(" ")[0],))
                    conn1.commit()
                    for item in list_material.keys():
                        if str(item[1]) == item_button.text.split(" ")[0]:
                            sql = "UPDATE material SET col=? WHERE address=? and ID = ?"
                            cursor1 = conn1.cursor()
                            cursor1.execute(sql, (str(list_material[item]), str(selff.button_text), item_button.text.split(" ")[0],))
                            conn1.commit()
                            self.keys_work[int(item_button.text.split(" ")[0])].text = str(list_material[item])
                    self.popup.dismiss()




                layout.add_widget(button_back)
                button_save.bind(on_press=okey_block)
                layout.add_widget(button_save)
                self.popup.open()

            if len(info_btn[2]) >= 27:
                text_1 = info_btn[2][:27]+'\n'+info_btn[2][27:]
                font = 40
            else:
                text_1 = info_btn[2]
                font = 40

            self.btn = Button(text=f'{info_btn[1]} {text_1}',
                              size_hint_y=None, size_hint_x=.7,
                              height=dp(60),
                              background_normal='button.png',
                              font_size=font,
                              color=[0, 0, 1, 1],
                              on_press = block_func)


            def move(button):
                if button.text == '+':
                    keys_move[button].text = str(int(keys_move[button].text) + 1)
                    keys_move[button].background_color = (0, 0, 1, 0.1)
                else:
                    if keys_move[button].text != '0':
                        keys_move[button].text = str(int(keys_move[button].text) - 1)
                        if keys_move[button].text == '0':
                            keys_move[button].background_color = (1, 0, 0, 0.1)

                    else:
                        keys_move[button].background_color = (1, 0, 0, 0.1)


            self.plus = Button(text=f'+',
                               size_hint_y=None, size_hint_x=.15,
                               height=dp(60),
                               background_normal='button.png',
                               font_size=40,
                               color=[0, 0, 1, 1],
                               on_press=lambda *args: move(*args))
            self.min = Button(text=f'-',
                              size_hint_y=None, size_hint_x=.15,
                              height=dp(60),
                              background_normal='button.png',
                              font_size=40,
                              color=[0, 0, 1, 1],
                              on_press=lambda *args: move(*args))
            self.col_work = TextInput(
                text=f'{list_total[info_btn][0]}',
                multiline=False,
                height=dp(30),
                pos_hint={'center_x': .5, 'center_y': .65},
                size_hint_y=.3,
                size_hint_x=.1,
                foreground_color=(0, 0, 1, 1),
                background_color=[0, 0, 1, 0.1],
                background_normal='',
                hint_text_color=(0, 0, 1, 1))

            if self.col_work.text == '0':
                self.col_work.background_color = (1, 0, 0, 0.1)
            if list_total[info_btn][1] == 't.y':
                self.btn.background_color = (1, 0, 0, 1)
            keys_move[self.plus] = self.col_work
            keys_move[self.min] = self.col_work
            self.keys_work[info_btn[1]] = self.col_work
            selff.layout.add_widget(self.btn)
            selff.layout.add_widget(self.plus)
            selff.layout.add_widget(self.col_work)
            selff.layout.add_widget(self.min)
            # work_list.append(info_btn.split("*")[0])
            # col_work_list.append(int(info_btn.split("*")[1]))
        def save():
            conn1 = sqlite3.connect(selff.roomsdatabase)
            cursor1 = conn1.cursor()
            cursor1.execute(sqlite_update_query, (str(selff.button_text),))
            conn1.commit()
            p=0
            for key in list_total.keys():
                if self.keys_work[int(key[1])].text == '':
                    p=1
                for v in self.keys_work[int(key[1])].text:
                    if (v in ['0','1','2','3','4','5','6','7','8','9'])==False:
                        p=2
                if p ==1:
                    break

                if str(list_total[key][0]) == self.keys_work[int(key[1])].text and str(list_total[key][1]) == 'f':
                    block1 = 'f'
                else:
                    block1 = 't.n'

                sql = f"""
                            INSERT INTO material (ID,address,col,block)
                            VALUES (?, ?, ?, ?);
                            """


                conn1 = sqlite3.connect(selff.roomsdatabase)
                cursor1 = conn1.cursor()
                cursor1.execute(sql,(str(key[1]), str(selff.button_text), self.keys_work[int(key[1])].text, block1,) )
                conn1.commit()
                cursor1.execute(
                    f"SELECT rowid, * FROM material WHERE address={str(selff.button_text)} and ID={str(key[1])}")
            if p!=0:
                layout = Screen()
                button_save = Button(text='Okeй', size_hint=(.95, .9), pos_hint={'center_x': .5, 'center_y': .5})
                if p==1:
                    self.popup = Popup(title=f'          Заполнены не все поля',
                                       content=layout, size_hint=(.7, .25), pos_hint={'center_x': .5, 'center_y': .65})
                else:
                    self.popup = Popup(title=f'          Введены лишние символы',
                                       content=layout, size_hint=(.7, .25), pos_hint={'center_x': .5, 'center_y': .65})

                def okt(a):
                    self.popup.dismiss()

                button_save.bind(on_press=okt)
                layout.add_widget(button_save)
                self.popup.open()
                return
            def ok():
                conn = sqlite3.connect(selff.mydatabase)
                cursor = conn.cursor()
                sql = "SELECT rowid, * FROM projects WHERE rowid=?"
                cursor.execute(sql, (selff.button_text,))
                selff.func_info(cursor.fetchall())
            ok()
        selff.button_add.on_press = lambda: save()
        selff.button_add.text = 'Сохранить'
        selff.sr.add_widget(root)

        self.windows = '2'

    def items_add_room(self,selff,j,input_text=''):
        def back_func():
                selff.add.text = '+'
                self.items_open_room(selff, selff.button_text)
        selff.back.on_press = back_func
        selff.sr.clear_widgets()
        def search_func():
            self.items_add_room(selff,j,input_text=self.search.text)
        selff.add.text = 'поиск'
        selff.add.font_size=30
        selff.add.on_press = search_func
        self.search = TextInput(
            text=f'{input_text}',
            multiline=False,
            height=dp(35),
            pos_hint={'center_x': .475, 'center_y': .875},
            size_hint_y=.05,
            size_hint_x=.75,
            foreground_color=(0, 0, 1, 1),
            background_color=[1, 1, 1],
            background_normal='',
            hint_text_color=(0, 0, 1, 1))
        selff.title.text = ''
        selff.sr.size_hint = (1, .95)
        selff.sr.pos_hint = {'center_x': .5, 'center_y': .58}
        layout = GridLayout(cols=2, spacing=1, size_hint=(1, None),
                            pos_hint={'center_x': .5, 'center_y': .5})
        layout.bind(minimum_height=layout.setter('height'))
        root = RecycleView(size_hint=(1, .8))
        root.add_widget(layout)
        selff.sr.add_widget(self.search)
        sql = "SELECT * FROM albums "
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        info_items = cursor1.fetchall()
        sql = "SELECT rowid, * FROM material WHERE address=?"
        conn1 = sqlite3.connect(selff.roomsdatabase)
        cursor1 = conn1.cursor()
        cursor1.execute(sql, (str(selff.button_text),))
        information_col_material = cursor1.fetchall()
        keys_move = {}
        i = 0
        for info_btn in info_items:
            if input_text.lower() in info_btn[1].lower():
                i += 1
                text = info_btn[1]
                btn = Button(text=f'{i}.{text}',
                             size_hint_y=None, size_hint_x=1,
                             height=dp(40),
                             background_normal='button.png',
                             font_size=20,
                             color=[0, 0, 1, 1])
                layout.add_widget(btn)
                active = CheckBox(active=False, size_hint_x=.1)
                for item in information_col_material:
                    if str(item[1]) == str(info_btn[0]):
                        active.active = True
                layout.add_widget(active)
                keys_move.setdefault(active, str(info_btn[0]))
        selff.sr.add_widget(root)
        def save(g):
            for key in keys_move.keys():
                okey = False
                if key.active == True:
                    for item in information_col_material:
                        if str(item[1]) == str(keys_move[key]):
                            okey = True
                    if okey == False:
                        sql = f"""
                                            INSERT INTO material (ID,address,col,block)
                                            VALUES (?, ?, ?, ?);
                                            """
                        conn1 = sqlite3.connect(selff.roomsdatabase)
                        cursor1 = conn1.cursor()
                        cursor1.execute(sql, (str(keys_move[key]), str(selff.button_text), '1', 't.y',))
                        conn1.commit()

            self.items_open_room( selff, j)


        button_add = Button(
            text='Сохранить',
            background_color=[0, 0, 1, 100],
            background_normal='',
            font_size=80,
            color=[1, 1, 1, 1],
            size_hint=(1, .1),
            pos_hint={'center_x': .5, 'center_y': -.06},
            on_press=save)
        selff.sr.add_widget(button_add)

    def open_group(self, selff,button_name):
        if button_name == 'Редактор материала':
            selff.title.text = 'Настройки'
        else:
            selff.title.text = button_name.split(' ')[1]
        selff.title.font_size = 80
        selff.back1.text = '<'
        selff.add.text = '+'
        selff.sr.clear_widgets()
        selff.sr.size_hint = (1, .95)
        selff.sr.pos_hint = {'center_x': .5, 'center_y': .58}
        layout = GridLayout(cols=2, spacing=10, size_hint=(1, None),
                            pos_hint={'center_x': .5, 'center_y': .5})
        layout.bind(minimum_height=layout.setter('height'))
        root = RecycleView(size_hint=(1, .8))
        root.add_widget(layout)
        sql = "SELECT rowid, * FROM name_group"
        conn1 = sqlite3.connect('items.db')
        cursor1 = conn1.cursor()
        cursor1.execute(sql)
        info = cursor1.fetchall()
        keys_move = {}
        i = 0
        for info_btn in info:

            if button_name == 'Редактор материала':
                if len(info_btn[1].split('.')) != 1:
                    continue
            else:
                sql = f"SELECT rowid, * FROM name_group WHERE id = '{button_name.split(' ')[0]}'"
                conn1 = sqlite3.connect('items.db')
                cursor1 = conn1.cursor()
                cursor1.execute(sql)
                if not(info_btn[1] in cursor1.fetchall()[0][3].split(',')):
                    continue
            i += 1
            if len(info_btn[2]) >= 29:
                text = info_btn[2][:28] + '\n' + info_btn[2][28:]
            else:
                text = info_btn[2]
            if info_btn[3] == None or len(info_btn[3].split(',')[0].split('.'))==1:
                btn = Button(text=f'{info_btn[1]} {text}',
                             size_hint_y=None, size_hint_x=.8,
                             height=dp(40),
                             background_normal='button.png',
                             font_size=20,
                             color=[0, 0, 1, 1],
                             on_press=lambda a: selff.func_open_group(a.text, keys_move))
            else:
                btn = Button(text=f'{info_btn[1]} {text}',
                             size_hint_y=None, size_hint_x=.8,
                             height=dp(40),
                             background_normal='button.png',
                             font_size=20,
                             color=[0, 0, 1, 1],
                             on_press=lambda a: selff.update_list_group(a.text, keys_move))


            layout.add_widget(btn)
            def delete(button):
                if keys_move[button]:
                    sqlite_update_query = """DELETE FROM name_group WHERE id = ?"""
                    conn1 = sqlite3.connect('items.db')
                    cursor1 = conn1.cursor()
                    cursor1.execute(sqlite_update_query, (str(selff.button_text),))
                    conn1.commit()
                selff.update_list_group(button_name, keys_move)
            redact = Button(text=f'Удалить',
                            size_hint_y=None, size_hint_x=.15,
                            height=dp(40),
                            background_normal='button.png',
                            font_size=20,
                            color=[0, 0, 1, 1],
                            on_press=lambda a: selff.update_list_group(a.text, keys_move))
            layout.add_widget(redact)
            keys_move.setdefault(redact, [info_btn[1]])
        selff.sr.add_widget(root)

        def add_group():
            layout = Screen()
            conn = sqlite3.connect("items.db")
            cursor = conn.cursor()
            font = Window.size[0] * Window.size[1] * .000025
            self.popup = Popup(title='     Создание',
                               content=layout, size_hint=(.8, .55))
            button_back = Button(text='Отменить', size_hint=(.5, .2), pos_hint={'center_x': .25, 'center_y': .1},
                                 font_size=font)
            button_save = Button(text='Сохранить', size_hint=(.5, .2), pos_hint={'center_x': .75, 'center_y': .1},
                                 font_size=font)
            input_name = TextInput(text=f'', size_hint=(.95, .16),
                                   pos_hint={'center_x': .5, 'center_y': .85},
                                   font_size=font)

            def check_2(a, b):
                if b.active == True and a.active == True:
                    b.active = False
                if b.active == False and a.active == False:
                    a.active = True

            kir = CheckBox(active=True, size_hint=(.1, .1), pos_hint={'center_x': .35, 'center_y': .65},
                           on_press=lambda a: check_2(kir, but))
            but = CheckBox(active=False, size_hint=(.1, .1), pos_hint={'center_x': .85, 'center_y': .65},
                           on_press=lambda a: check_2(but, kir))
            info_list = [['Материал', 'Подгруппа']]
            y_x = .65
            for i in info_list:
                label1 = Label(text=i[0], size_hint=(.25, .15), pos_hint={'center_x': .15, 'center_y': y_x},
                               font_size=font * .8)
                label2 = Label(text=i[1], size_hint=(.25, .15), pos_hint={'center_x': .65, 'center_y': y_x},
                               font_size=font * .8)
                layout.add_widget(label2)
                layout.add_widget(label1)
                y_x -= .15

            def ok_save(*args):
                sql = f"SELECT rowid, * FROM name_group WHERE id = '{button_name.split(' ')[0]}'"
                conn1 = sqlite3.connect('items.db')
                cursor1 = conn1.cursor()
                cursor1.execute(sql)
                info_g = cursor1.fetchall()[0]
                inter = 1
                for list_group in info_g[3].split(','):
                    if list_group.split('.')[len(list_group.split('.'))-1] != None and list_group.split('.')[len(list_group.split('.'))-1] != '':
                        if inter <= int(list_group.split('.')[len(list_group.split('.'))-1]):
                            inter=int(list_group.split('.')[len(list_group.split('.'))-1])+1


                sql = f"""
                    INSERT INTO name_group (id,name,info)
                    VALUES ('{button_name.split(' ')[0]+'.'+str(inter)}','{input_name.text}', '');
                    """
                cursor.execute(sql)
                conn.commit()

                sql = F"UPDATE name_group SET info='{info_g[3]+','+button_name.split(' ')[0]+'.'+str(inter)}' WHERE id = '{button_name.split(' ')[0]}'"
                cursor.execute(sql)
                conn.commit()

                self.popup.dismiss()
                selff.update_list_group(button_name.split(' ')[0]+'.'+str(inter)+' '+input_name.text,key=' ')

            button_back.bind(on_press=lambda *args: self.popup.dismiss())
            layout.add_widget(button_back)
            button_save.bind(on_press=lambda *args: ok_save())
            layout.add_widget(button_save)
            layout.add_widget(input_name)
            layout.add_widget(kir)
            layout.add_widget(but)
            self.popup.open()

        def okey():
            sql = f"SELECT rowid, * FROM name_group "
            conn1 = sqlite3.connect('items.db')
            cursor1 = conn1.cursor()
            cursor1.execute(sql)
            inf = cursor1.fetchall()
            for info_info in inf:
                if button_name == 'Редактор материала':
                    selff.okey()
                    return
                elif info_info[3] != None and len(button_name.split(' ')[0].split('.')) ==1:
                    nfo = 'Редактор материала'
                    selff.update_list_group(nfo, button_name)
                elif info_info[3] != None and button_name.split(' ')[0] in info_info[3].split(','):
                    nfo = info_info[1] + ' ' + info_info[2]
                    selff.update_list_group(nfo, button_name)

        selff.back1.on_press = okey
        selff.add.on_press = add_group
